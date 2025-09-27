"""
Excel report generation service.

Serwis generowania raportów Excel.
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import json


class ExcelGenerator:
    """Excel report generator using openpyxl."""

    def __init__(self):
        self.workbook = None
        self.worksheet = None

    def generate_report_excel(self, report_data: Dict[str, Any], sections: List[Dict[str, Any]], output_path: Path) -> Path:
        """Generate a complete Excel report."""

        self.workbook = Workbook()

        # Remove default worksheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

        # Create summary sheet
        self._create_summary_sheet(report_data)

        # Create sections as separate worksheets
        for section in sections:
            self._create_section_sheet(section)

        # Create charts sheet
        self._create_charts_sheet(sections)

        # Save workbook
        self.workbook.save(output_path)
        return output_path

    def _create_summary_sheet(self, report_data: Dict[str, Any]) -> None:
        """Create the summary worksheet."""

        ws = self.workbook.create_sheet("Summary", 0)

        # Title
        ws['A1'] = report_data.get('title', 'System Report')
        ws['A1'].font = Font(size=18, bold=True, color='1F2937')

        # Report metadata
        metadata = [
            ('Report Type:', report_data.get('report_type', '').replace('_', ' ').title()),
            ('Generated:', datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')),
            ('Report ID:', report_data.get('id', 'N/A')),
            ('Format:', report_data.get('format', 'Excel').upper()),
            ('Status:', report_data.get('status', 'completed').title())
        ]

        if report_data.get('description'):
            metadata.append(('Description:', report_data['description']))

        if report_data.get('date_range'):
            date_range = report_data['date_range']
            metadata.extend([
                ('Start Date:', date_range.get('start_date', 'N/A')),
                ('End Date:', date_range.get('end_date', 'N/A'))
            ])

        # Add metadata to sheet
        for i, (key, value) in enumerate(metadata, start=3):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True, color='6B7280')
            ws[f'B{i}'].font = Font(color='111827')

        # Style the sheet
        self._apply_summary_styling(ws, len(metadata) + 2)

    def _create_section_sheet(self, section: Dict[str, Any]) -> None:
        """Create a worksheet for a report section."""

        section_name = section.get('name', 'Section').replace('_', ' ').title()
        # Excel sheet names have a 31 character limit
        sheet_name = section_name[:31] if len(section_name) > 31 else section_name

        ws = self.workbook.create_sheet(sheet_name)

        # Section title
        ws['A1'] = section_name
        ws['A1'].font = Font(size=14, bold=True, color='1F2937')

        current_row = 3

        # Section data
        data = section.get('data', {})
        if data:
            # Create data table
            ws[f'A{current_row}'] = 'Metric'
            ws[f'B{current_row}'] = 'Value'
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'].font = Font(bold=True)

            current_row += 1

            for key, value in data.items():
                metric_name = key.replace('_', ' ').title()

                ws[f'A{current_row}'] = metric_name

                if isinstance(value, float):
                    if 'percentage' in key.lower() or 'rate' in key.lower():
                        ws[f'B{current_row}'] = value / 100  # Excel percentage
                        ws[f'B{current_row}'].number_format = '0.00%'
                    else:
                        ws[f'B{current_row}'] = value
                        ws[f'B{current_row}'].number_format = '#,##0.00'
                elif isinstance(value, int):
                    ws[f'B{current_row}'] = value
                    ws[f'B{current_row}'].number_format = '#,##0'
                elif isinstance(value, dict) or isinstance(value, list):
                    ws[f'B{current_row}'] = json.dumps(value, indent=2) if len(str(value)) < 1000 else str(value)[:1000] + "..."
                else:
                    ws[f'B{current_row}'] = str(value)

                current_row += 1

            # Style the data table
            self._apply_table_styling(ws, 3, current_row - 1, 2)

        current_row += 2

        # Chart data
        charts = section.get('charts', [])
        if charts:
            ws[f'A{current_row}'] = 'Chart Data'
            ws[f'A{current_row}'].font = Font(bold=True, color='1F2937')
            current_row += 2

            for chart in charts:
                chart_title = chart.get('title', 'Chart')
                chart_data = chart.get('data', [])
                chart_labels = chart.get('labels', [])

                if chart_data and chart_labels:
                    # Chart title
                    ws[f'A{current_row}'] = chart_title
                    ws[f'A{current_row}'].font = Font(bold=True)
                    current_row += 1

                    # Chart data table
                    ws[f'A{current_row}'] = 'Label'
                    ws[f'B{current_row}'] = 'Value'
                    current_row += 1

                    for label, value in zip(chart_labels, chart_data):
                        ws[f'A{current_row}'] = label
                        ws[f'B{current_row}'] = value
                        if isinstance(value, (int, float)):
                            ws[f'B{current_row}'].number_format = '#,##0.00'
                        current_row += 1

                    current_row += 1

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_charts_sheet(self, sections: List[Dict[str, Any]]) -> None:
        """Create a worksheet with charts."""

        ws = self.workbook.create_sheet("Charts")

        ws['A1'] = 'Report Charts'
        ws['A1'].font = Font(size=14, bold=True, color='1F2937')

        current_row = 3
        chart_position_row = 3

        for section in sections:
            charts = section.get('charts', [])

            for chart in charts:
                chart_type = chart.get('type', 'bar')
                chart_title = chart.get('title', 'Chart')
                chart_data = chart.get('data', [])
                chart_labels = chart.get('labels', [])

                if not chart_data or not chart_labels:
                    continue

                # Create data for chart
                ws[f'A{current_row}'] = chart_title
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1

                # Headers
                ws[f'A{current_row}'] = 'Category'
                ws[f'B{current_row}'] = 'Value'
                start_row = current_row
                current_row += 1

                # Data
                for label, value in zip(chart_labels, chart_data):
                    ws[f'A{current_row}'] = label
                    ws[f'B{current_row}'] = value
                    current_row += 1

                end_row = current_row - 1

                # Create chart
                try:
                    if chart_type == 'bar':
                        chart_obj = BarChart()
                    elif chart_type == 'line':
                        chart_obj = LineChart()
                    elif chart_type == 'pie':
                        chart_obj = PieChart()
                    else:
                        chart_obj = BarChart()

                    # Add data to chart
                    data_ref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=end_row, max_col=2)
                    cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=end_row)

                    chart_obj.add_data(data_ref, titles_from_data=False)
                    chart_obj.set_categories(cats_ref)
                    chart_obj.title = chart_title

                    # Position chart
                    chart_obj.anchor = f'D{chart_position_row}'
                    chart_obj.width = 15
                    chart_obj.height = 10

                    ws.add_chart(chart_obj)

                    chart_position_row += 20

                except Exception as e:
                    # If chart creation fails, just add a note
                    ws[f'D{chart_position_row}'] = f"Chart creation failed: {str(e)}"
                    chart_position_row += 5

                current_row += 2

    def _apply_summary_styling(self, ws, last_row: int) -> None:
        """Apply styling to summary sheet."""

        # Title styling
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

        # Header fill
        header_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

        # Apply border to metadata area
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(3, last_row + 1):
            for col in ['A', 'B']:
                cell = ws[f'{col}{row}']
                cell.border = thin_border
                if col == 'A':
                    cell.fill = header_fill

    def _apply_table_styling(self, ws, start_row: int, end_row: int, num_cols: int) -> None:
        """Apply styling to data tables."""

        # Header styling
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        # Data styling
        alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')

                if row == start_row:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                elif row % 2 == 0:  # Alternate row coloring
                    cell.fill = alt_fill